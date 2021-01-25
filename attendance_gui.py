import os
import csv
import smtplib
import datetime
import openpyxl
import threading
from tkinter import *
from email import encoders
from tkinter import messagebox
import paho.mqtt.client as mqtt
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart


class AttendanceSystemGUI:
    clientName = "Attendance-Station"
    mosquitoBrokerIP = "127.0.0.1"
    mosquitoBrokerPort = 1883
    mainWindowTitle = "Wireless Student Attendance System"
    courseCode = ""
    attendanceFolder = ""
    attendanceDate = datetime.datetime.now()
    listOfAttendees = {}
    listOfNamesAndID = {}

    def __init__(self):
        """DOCUMENTATION GOES HERE"""
        self.mainWindow = Tk()
        self.mosquitoClient = mqtt.Client(self.clientName)

        frmTittle = Frame()
        frmTittle.pack(side=TOP)
        picTittle = PhotoImage(file="logo.png")
        lblTittle = Label(frmTittle, image=picTittle)
        lblTittle.image = picTittle
        lblTittle.pack(side=TOP)

        frmCourse = Frame(height=10)
        frmCourse.pack(side=TOP)
        Label(frmCourse, text="Course Code").pack(side=LEFT)
        self.eCourse = Entry(frmCourse, bd=6, relief=RIDGE,
                             width=40, justify=CENTER)
        self.eCourse.bind("<Return>", self.checkAttendanceFile)
        self.eCourse.pack(side=LEFT, padx=10, pady=1)
        self.eCourse.focus()

        frmStatus = Frame(height=15)
        frmStatus.pack(side=TOP)
        picStart = PhotoImage(file="start.png")
        self.btnStart = Button(frmStatus, relief=FLAT, image=picStart,
                               state=DISABLED, command=lambda:
            [threading.Thread(target=self.startAttendance, daemon=True).start(), self.updateButtons(1)])
        self.btnStart.config
        self.btnStart.image = picStart
        self.btnStart.pack(side=LEFT)
        picStop = PhotoImage(file="stop.png")
        self.btnStop = Button(frmStatus, relief=FLAT, image=picStop,
                              state=DISABLED, command=self.stopAttendance)
        self.btnStop.image = picStop
        self.btnStop.pack(side=LEFT)
        picExport = PhotoImage(file="export.png")
        self.btnExport = Button(frmStatus, relief=FLAT, image=picExport,
                                state=DISABLED, command=self.updateAttendance)
        self.btnExport.image = picExport
        self.btnExport.pack(side=LEFT)
        picEmail = PhotoImage(file="email.png")
        self.btnEmail = Button(frmStatus, relief=FLAT, image=picEmail,
                               state=DISABLED, command=self.emailAttendance)
        self.btnEmail.image = picEmail
        self.btnEmail.pack(side=LEFT)

        self.mainWindow.resizable(False, False)
        self.mainWindow.title(self.mainWindowTitle)
        self.mainWindow.geometry(
            '+' + str(int(self.mainWindow.winfo_screenwidth() / 2) -
                      int(self.mainWindow.winfo_screenwidth() / 6)) +
            '+' + str(int(self.mainWindow.winfo_screenheight() / 2) -
                      int(self.mainWindow.winfo_screenheight() / 6)))
        self.mainWindow.protocol("WM_DELETE_WINDOW", self.exitProgram)
        self.mainWindow.mainloop()

    def updateButtons(self, whichButton):
        if whichButton == 1:
            self.btnStart.config(state=DISABLED)
            self.btnStop.config(state=ACTIVE)
            self.btnExport.config(state=DISABLED)
            self.btnEmail.config(state=DISABLED)
            print("Started Taking Attendance")
        elif whichButton == 2:
            self.btnStart.config(state=ACTIVE)
            self.btnStop.config(state=DISABLED)
            self.btnExport.config(state=ACTIVE)
            self.btnEmail.config(state=DISABLED)
        elif whichButton == 3:
            self.btnEmail.config(state=ACTIVE)
        elif whichButton == 4:
            pass
        elif whichButton == 5:
            self.btnStart.config(state=ACTIVE)

    def startAttendance(self):
        """DOCUMENTATION GOES HERE"""
        """
        This function sets up and starts the Mosquito MQTT client connection
        """
        self.mosquitoClient.connect(self.mosquitoBrokerIP, port=self.mosquitoBrokerPort)
        self.mosquitoClient.on_message = self.readTagID
        self.mosquitoClient.subscribe("ATTENDANCE")
        self.mosquitoClient.publish("STATION", "1")
        self.mosquitoClient.loop_forever()

    def stopAttendance(self):
        """DOCUMENTATION GOES HERE"""
        """
        This function stops the Mosquito MQTT client connection
        """
        self.updateButtons(2)
        self.mosquitoClient.publish("STATION", "0")
        self.mosquitoClient.unsubscribe("ATTENDANCE")
        self.mosquitoClient.loop_stop()
        self.mosquitoClient.disconnect()
        print("Stooped Taking Attendance")

    def updateAttendance(self):
        """DOCUMENTATION GOES HERE"""
        """
        This function either creates or updates the excel file with attendance information
        """
        self.updateButtons(3)
        if not os.path.isdir(self.attendanceFolder[0:6]):
            os.mkdir(self.attendanceFolder[0:6])
        with open(self.attendanceFolder + self.courseCode + '_Student_Names.csv', 'w', newline='') as csvfile:
            fieldnames = ['Id', 'Name']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            for items in self.listOfNamesAndID.keys():
                writer.writerow({'Id': items, 'Name': self.listOfNamesAndID[items]})
        print("Name and ID file Created")

        if not os.path.isfile(self.attendanceFolder + self.courseCode + '_Attendance.xlsx'):
            attendanceFileWorkBook = openpyxl.Workbook(self.attendanceFolder +
                                                       self.courseCode + '_Attendance.xlsx')
            attendanceFileSheet = attendanceFileWorkBook.create_sheet('Attendance')
            attendanceFileSheet.append(["ID", "Name", str(self.attendanceDate.date())])
            for records in self.listOfAttendees.keys():
                attendanceFileSheet.append([records, self.listOfAttendees[records], "P"])
        else:
            attendanceFileWorkBook = openpyxl.load_workbook(self.attendanceFolder +
                                                            self.courseCode + '_Attendance.xlsx')
            attendanceFileSheet = attendanceFileWorkBook.active
            newColumn = attendanceFileSheet.max_column+1
            newRow = attendanceFileSheet.max_row+1
            attendanceFileSheet.cell(1, newColumn, str(self.attendanceDate.date()))
            for col in attendanceFileSheet.iter_cols(1, 1, 2):
                for records in self.listOfAttendees.keys():
                    isNewStudent = True
                    for currentCell in col:
                        if currentCell.value == records:
                            isNewStudent = False
                            attendanceFileSheet.cell(currentCell.row, newColumn, "P")
                    if isNewStudent:
                        attendanceFileSheet.cell(newRow, 1, records)
                        attendanceFileSheet.cell(newRow, 2, self.listOfAttendees[records])
                        attendanceFileSheet.cell(newRow, newColumn, "P")
        attendanceFileWorkBook.save(self.attendanceFolder + self.courseCode + '_Attendance.xlsx')
        print("Attendance file Created")
        messagebox.showinfo("Export File", "The attendance file has been created/updated successfully!")

    def emailAttendance(self):
        """DOCUMENTATION GOES HERE"""
        def sendEmail(emailFrom, emailPassword, emailTo, emailAttachment, emailSubject = "", emailBody = ""):
            """DOCUMENTATION GOES HERE"""
            emailWindow.update()
            try:
                msg = MIMEMultipart()
                msg['From'] = emailFrom
                msg['To'] = emailTo
                msg['Subject'] = emailSubject
                msg.attach(MIMEText(emailBody, 'plain'))
                attachment = open(emailAttachment, "rb")
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', "attachment; filename= %s" % emailAttachment.split("/")[1])
                msg.attach(part)
                server = smtplib.SMTP('outlook.office365.com', 587)
                server.starttls()
                server.login(emailFrom, emailPassword)
                text = msg.as_string()
                server.sendmail(emailFrom, emailTo, text)
                server.quit()
            except:
                print("Not Emailed!!!!!")
                messagebox.showerror("Error", "Email could not be sent! Please try again.")
                return False
            else:
                print("Emailed")
                messagebox.showinfo("Success", "Email has been sent to " + eTo.get())
                emailWindow.destroy()
                return True

        self.updateButtons(4)
        emailWindow = Toplevel()
        emailWindow.resizable(False, False)
        emailWindow.title("Email Attendance")
        emailWindow.grab_set()
        emailWindow.geometry(
            '+' + str(int(emailWindow.winfo_screenwidth() / 2) -
                      int(emailWindow.winfo_screenwidth() / 9)) +
            '+' + str(int(emailWindow.winfo_screenheight() / 2) -
                      int(emailWindow.winfo_screenheight() / 14)))
        Label(emailWindow, text="Email Information", font=("Courier", 16), fg="red", bg="black").pack(fill=BOTH)

        frmFrom = Frame(emailWindow)
        frmFrom.pack(side=TOP)
        Label(frmFrom, text="        From").pack(side=LEFT)
        eFrom = Entry(frmFrom, bd=6, relief=RIDGE, width=40)
        eFrom.pack(side=LEFT, padx=10, pady=1)

        frmPassword = Frame(emailWindow)
        frmPassword.pack(side=TOP)
        Label(frmPassword, text="Password").pack(side=LEFT)
        ePassword = Entry(frmPassword, bd=6, relief=RIDGE, show="*", width=40)
        ePassword.pack(side=LEFT, padx=10, pady=1)

        frmTo = Frame(emailWindow)
        frmTo.pack(side=TOP)
        Label(frmTo, text="            To").pack(side=LEFT)
        eTo = Entry(frmTo, bd=6, relief=RIDGE, width=40)
        eTo.pack(side=LEFT, padx=10, pady=1)

        frmAttachment = Frame(emailWindow)
        frmAttachment.pack(side=TOP)
        Label(frmAttachment, text="Attachment").pack(side=LEFT)
        Label(frmAttachment, text=self.courseCode + "_Attendance.xlsx", bd=6,
              bg="#3fccb5", relief=RIDGE, width=32).pack(side=LEFT, padx=10, pady=1)

        frmSubject = Frame(emailWindow)
        frmSubject.pack(side=TOP)
        Label(frmSubject, text="   Subject").pack(side=LEFT)
        eSubject = Entry(frmSubject, bd=6, relief=RIDGE, width=40)
        eSubject.pack(side=LEFT, padx=10, pady=1)

        picSendEmail = PhotoImage(file="email.png")
        btnSendEmail = Button(emailWindow, image=picSendEmail,
                              command=lambda: sendEmail(eFrom.get(), ePassword.get(),
                                                        eTo.get(), self.attendanceFolder + self.courseCode +
                                                        "_Attendance.xlsx", eSubject.get()))
        btnSendEmail.image = picSendEmail
        btnSendEmail.pack()
        emailWindow.update()
        self.mainWindow.update()

    def checkAttendanceFile(self, event=None):
        """DOCUMENTATION GOES HERE"""
        self.courseCode = str(self.eCourse.get()).upper()
        self.eCourse.config(text=self.courseCode)
        if not (len(self.courseCode) == 9) or not (self.courseCode[0:2].isalpha() and
                                                   self.courseCode[3:5].isdigit() and
                                                   self.courseCode[6:].isalpha()):
            self.eCourse.config(bg="red", fg="white")
            self.courseCode = ""
        else:
            self.eCourse.config(bg="green", fg="white")
            self.attendanceFolder = self.courseCode[0:6] + '/'
            # self.updateAttendance()
            if not os.path.isdir(self.attendanceFolder[0:6]):
                os.mkdir(self.attendanceFolder[0:6])
            if os.path.isfile(self.attendanceFolder + self.courseCode + '_Student_Names.csv'):
                with open(self.attendanceFolder + self.courseCode + '_Student_Names.csv', newline='') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        self.listOfNamesAndID[row['Id']] = row['Name']
            print(self.listOfNamesAndID)
            self.updateButtons(5)

    def readTagID(self, client, data, uid):
        """DOCUMENTATION GOES HERE"""
        def newStudent():
            """DOCUMENTATION GOES HERE"""
            def addStudent(event=None):
                """DOCUMENTATION GOES HERE"""
                if eName.get():
                    self.listOfAttendees[tagID] = eName.get()
                    self.listOfNamesAndID[tagID] = eName.get()
                    print(self.listOfAttendees)
                    self.mosquitoClient.publish("STATION", "S")
                    newStudentWindow.destroy()

            newStudentWindow = Toplevel()
            newStudentWindow.resizable(False, False)
            newStudentWindow.title("New Student")
            newStudentWindow.grab_set()
            newStudentWindow.geometry(
                '+' + str(int(newStudentWindow.winfo_screenwidth() / 2) -
                          int(newStudentWindow.winfo_screenwidth() / 9)) +
                '+' + str(int(newStudentWindow.winfo_screenheight() / 2) -
                          int(newStudentWindow.winfo_screenheight() / 14)))
            Label(newStudentWindow, text="Add New", font=("Courier", 16), fg="red", bg="black").pack(fill=BOTH)
            frmName = Frame(newStudentWindow)
            frmName.pack(side=TOP)
            Label(frmName, text="Student Name").pack(side=LEFT)
            eName = Entry(frmName, bd=6, relief=RIDGE, width=40)
            eName.bind("<Return>", addStudent)
            eName.pack(side=LEFT, padx=10, pady=1)

            picAddStudent = PhotoImage(file="add.png")
            btnAdd = Button(newStudentWindow, image=picAddStudent, command=addStudent)
            btnAdd.image = picAddStudent
            btnAdd.pack(fill=BOTH)
            eName.focus()

        tagID = str(uid.payload.decode("utf-8"))
        if tagID == "1":
            self.stopAttendance()
        elif tagID not in self.listOfAttendees:
            if tagID not in self.listOfNamesAndID:
                newStudent()
            else:
                self.listOfAttendees[tagID] = self.listOfNamesAndID[tagID]
                print(self.listOfAttendees)
                self.mosquitoClient.publish("STATION", "S")
        else:
            self.mosquitoClient.publish("STATION", "A")
            print(self.listOfAttendees)

    def exitProgram(self):
        """DOCUMENTATION GOES HERE"""
        self.stopAttendance()
        self.mainWindow.destroy()


attendanceSystem = AttendanceSystemGUI()
